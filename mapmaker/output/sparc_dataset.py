# ===============================================================================
#
#  Flatmap viewer and annotation tools
#
#  Copyright (c) 2019 - 2023 David Brooks
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
# ===============================================================================

from datetime import datetime, timezone, timedelta
from pathlib import Path
from dataclasses import dataclass
from io import BytesIO
import json
import logging
import mimetypes
import shutil
from zipfile import ZipFile, ZipInfo, ZIP_DEFLATED
from typing import Optional

# ===============================================================================

import openpyxl
import requests

# ===============================================================================

from mapmaker.flatmap import FlatMap, Manifest
from mapmaker.utils import pathlib_path

# ===============================================================================

MAPPING_URL = "mapmaker/output/data_mapping.json"
PRIMARY = 0
DERIVATIVE = 1
GIT_RELEASE = "https://api.github.com/repos/SciCrunch/sparc-curation/git/refs/tags/{tag}"
GIT_TEMPLATE = "https://github.com/SciCrunch/sparc-curation/blob/{sha}/resources/DatasetTemplate/{file}.xlsx?raw=true"

# ===============================================================================


class VersionMapping:
    def __init__(self, version=None):
        with open(MAPPING_URL, 'r') as f:
            mappings = json.load(f)
            self.__mapping = None
            if version is None:
                self.__mapping = mappings[0]
            else:
                for v in mappings:
                    if v['version'] == version:
                        self.__mapping = v
            if self.__mapping is None:
                raise Exception(
                    f'Dataset-Description version-{version} is not available')

            # get SHA
            git_release = GIT_RELEASE.format(tag=self.__mapping['version'])
            response = requests.get(git_release, timeout=10).json()
            self.__sha = response['object']['sha']

    def get_mapping(self, other_params):
        """
        : other_params: is a dictionary containing other data such as uuid and version.
        """
        for m in self.__mapping['mapping']:
            if len(m[1]) > 0:
                param = m[1][-1]
                if param in other_params:
                    m[2] = other_params[param]
        return self.__mapping

    def __load_template_workbook(self, template_link):
        """
        : template_link: link to dataset_description.xlsx.
        """
        headers = {'Content-Type': 'application/xlsx'}
        template = requests.request('GET', template_link, headers=headers)
        workbook = openpyxl.load_workbook(BytesIO(template.content))
        return workbook

    @property
    def description_template(self):
        url = GIT_TEMPLATE.format(sha=self.__sha, file="dataset_description")
        return self.__load_template_workbook(url)

    @property
    def submission_template(self):
        url = GIT_TEMPLATE.format(sha=self.__sha, file="submission")
        return self.__load_template_workbook(url)

# ===============================================================================


class DatasetDescription:
    def __init__(self, flatmap, version):
        """
        : flatmap: is a Flatmap instance.
        : version: is SDS version.
        """
        other_params = {
            'id': ['URL', 'UUID'],
            'id_type': [flatmap.metadata.get('source'), flatmap.uuid]
            }

        version_mapping = VersionMapping(version)
        self.__mapping =version_mapping.get_mapping(other_params)
        self.__desc_workbook = version_mapping.description_template

    def write(self, description_file):
        if description_file.startswith('file'):
            description_file = pathlib_path(description_file)
        with open(description_file, 'r') as fd:
            self.__description = json.load(fd)
        for m in self.__mapping['mapping']:
            self.__write_cell(m)

    def __write_cell(self, map):
        worksheet = self.__desc_workbook.worksheets[0]
        data_pos = self.__mapping['data_pos']
        key, dsc, default = map
        values = default if isinstance(default, list) else [default]

        if len(dsc) == 1:
            if dsc[0] in self.__description:
                values = self.__description[dsc[-1]] if isinstance(
                    self.__description[dsc[-1]], list) else [self.__description[dsc[-1]]]
        elif len(dsc) > 1:
            tmp_values = self.__description
            for d in dsc:
                if isinstance(tmp_values, dict):
                    tmp_values = tmp_values.get(d, {})
                elif isinstance(tmp_values, list):
                    tmp_values = [val.get(d, '') for val in tmp_values]

            if len(tmp_values) > 0:
                values = tmp_values if isinstance(
                    tmp_values, list) else [tmp_values]

        for row in worksheet.rows:
            if row[0].value is None:
                break
            if row[0].value.lower().strip() == key:
                for pos in range(len(values)):
                    row[pos+data_pos].value = str(values[pos])

    def get_bytes(self):
        buffer = BytesIO()
        self.__desc_workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def get_json(self):
        return self.__description

    def close(self):
        self.__desc_workbook.close()

# ===============================================================================


@dataclass
class DatasetFile:
    filename: str
    fullpath: Path
    timestamp: datetime
    description: str
    file_type: str

# ===============================================================================


class DirectoryManifest:
    COLUMNS = (
        'filename',
        'timestamp',
        'description',
        'file type',
    )

    def __init__(self, metadata_columns: Optional[list[str]] = None):
        self.__metadata_columns = metadata_columns if metadata_columns is not None else []
        self.__files = []
        self.__file_records = []

    def __get_repo_datetime(self, fullpath):
        return datetime.fromtimestamp(fullpath.stat().st_mtime)

    @property
    def files(self):
        return self.__files

    @property
    def file_list(self):
        return [f.fullpath for f in self.__files]

    def add_file(self, filename, description, **metadata):
        fullpath = filename.resolve()
        file_type = mimetypes.guess_type(filename, strict=False)[0]
        file_type = fullpath.suffix if file_type is None else file_type
        dataset_file = DatasetFile(fullpath.name,
                                   fullpath,
                                   self.__get_repo_datetime(fullpath),
                                   description,
                                   file_type)
        self.__files.append(dataset_file)
        record: list[str | None] = [
            dataset_file.filename,
            dataset_file.timestamp.isoformat(),
            dataset_file.description,
            dataset_file.file_type
        ]
        for column_name in self.__metadata_columns:
            record.append(metadata.get(column_name))
        self.__file_records.append(record)

    def get_bytes(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for col, value in enumerate(self.COLUMNS + tuple(self.__metadata_columns), start=1):
            worksheet.cell(row=1, column=col, value=value)
        for row, record in enumerate(self.__file_records, start=2):
            for col, value in enumerate(record, start=1):
                worksheet.cell(row=row, column=col, value=value)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        workbook.close()
        return buffer

# ===============================================================================


class FlatmapSource:
    def __init__(self, manifest, flatmap, data_type=PRIMARY):
        """
        : manifest: a Manifest instance.
        : flatmap: a Flatmap instance.
        : data_type: a binary whether the source manage by repo or not.
        """

        # creating dataset manifest.xlsx
        species = manifest.models
        metadata = {'species': species} if species is not None else {}

        directory_manifest = DirectoryManifest(list(metadata.keys()))

        if data_type == PRIMARY:
            # adding files to be store in primary directory
            directory_manifest.add_file(pathlib_path(
                manifest.description), 'flatmap description', **metadata)
            if manifest.anatomical_map is not None:
                directory_manifest.add_file(pathlib_path(
                    manifest.anatomical_map), 'flatmap annatomical map', **metadata)
            if manifest.properties is not None:
                directory_manifest.add_file(pathlib_path(
                    manifest.properties), 'flatmap properties', **metadata)
            if manifest.connectivity_terms is not None:
                directory_manifest.add_file(pathlib_path(
                    manifest.connectivity_terms), 'flatmap connectivity terms', **metadata)
            for connectivity_file in manifest.connectivity:
                directory_manifest.add_file(pathlib_path(
                    connectivity_file), 'flatmap connectivity', **metadata)
            for source in manifest.sources:
                if source['href'].split(':', 1)[0] in ['file']:
                    directory_manifest.add_file(pathlib_path(
                        source['href']), 'flatmap source', **metadata)
            manifest_dir = pathlib_path(manifest.description).parent
            manifest_path = (
                manifest_dir / pathlib_path(manifest.url).name).resolve()
            directory_manifest.add_file(
                manifest_path, 'manifest to built map', **metadata)
        elif data_type == DERIVATIVE:
            
            for file in pathlib_path(flatmap.map_dir).glob('[!.]*'):
                if file.is_file():
                    directory_manifest.add_file(
                        file, 'derivative file to be used by map server', **metadata)

        self.__directory_manifests = [directory_manifest]

    @property
    def directory_manifests(self):
        return self.__directory_manifests

    @property
    def dataset_image(self):
        for directory_manifest in self.__directory_manifests:
            for file in directory_manifest.files:
                if file.filename.endswith('.svg'):
                    return file.fullpath
        return None

    def copy_to_archive(self, archive, target):
        for directory_manifest in self.directory_manifests:
            for file in directory_manifest.files:
                zinfo = ZipInfo.from_file(
                    str(file.fullpath), arcname=f'{target}/{file.filename}')
                zinfo.compress_type = ZIP_DEFLATED
                timestamp = file.timestamp
                zinfo.date_time = (timestamp.year, timestamp.month, timestamp.day,
                                   timestamp.hour, timestamp.minute, timestamp.second)
                with open(file.fullpath, "rb") as src, archive.open(zinfo, 'w') as dest:
                    shutil.copyfileobj(src, dest, 1024*8)
            manifest_workbook = directory_manifest.get_bytes()
            archive.writestr(f'{target}/manifest.xlsx',
                             manifest_workbook.getvalue())
            manifest_workbook.close()

# ===============================================================================


class SparcDataset:
    def __init__(self, manifest: Manifest, flatmap: FlatMap):
        self.__manifest = manifest
        self.__flatmap = flatmap

    def generate(self):
        # generate dataset_description
        self.__description = DatasetDescription(self.__flatmap, version=None)
        try:
            self.__description.write(self.__manifest.description)
        except Exception:
            logging.error(
                f'Cannot create dataset: Cannot open: {self.__manifest.description}')

        # generate primary source
        self.__primary = FlatmapSource(self.__manifest, self.__flatmap)

        # generate derivative source
        self.__derivative = FlatmapSource(
            self.__manifest, self.__flatmap, data_type=DERIVATIVE)

    def save(self, dataset: str):
        # create archive
        dataset_archive = ZipFile(dataset, mode='w', compression=ZIP_DEFLATED)

        # adding dataset_description
        desc_bytes = self.__description.get_bytes()
        dataset_archive.writestr(
            'files/dataset_description.xlsx', desc_bytes.getvalue())
        desc_bytes.close()
        self.__description.close()

        # copy primary data
        self.__primary.copy_to_archive(dataset_archive, 'files/primary')

        # this one save derivatives
        self.__derivative.copy_to_archive(dataset_archive, 'files/derivative')

        # create and save proper readme file, generated for dataset_description
        self.__add_readme(dataset_archive)

        # save banner
        banner_file = pathlib_path(self.__flatmap.full_filename(f'{self.__flatmap.id}.svg'))
        if banner_file.exists():
            dataset_archive.write(banner_file, 'files/banner.svg')

        # add submission file
        self.__add_submission(dataset_archive, 'files/submission.xlsx')

        # close archive
        dataset_archive.close()

    def __add_readme(self, archive):
        # load flatmap description
        readme = ['# FLATMAP DESCRIPTION'] + \
            self.__metadata_parser(self.__description.get_json())
        # load flatmat setup
        readme += ['# FLATMAP SETTINGS'] + \
            self.__metadata_parser(self.__flatmap.metadata)
        archive.writestr('files/readme.md', '\n'.join(readme))

    def __metadata_parser(self, data):
        metadata = []
        for key, val in data.items():
            metadata += [f'## {key.capitalize()}']
            if isinstance(val, dict):
                for subkey, subval in val.items():
                    metadata += [f'- {subkey}: {subval}']
            elif isinstance(val, list):
                for subval in val:
                    if isinstance(subval, dict):
                        for subsubkey, subsubval in subval.items():
                            metadata += [f'- {subsubkey}: {subsubval}']
                    else:
                        metadata += [f'- {subval}']
            else:
                metadata += [str(val)]
        return metadata

    def __add_submission(self, archive, filepath):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        columns = ['Submission Item', 'Definition', 'Value']
        header_fill = openpyxl.styles.PatternFill(
            start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        col_width = [30, 45, 30]
        for col, value in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=col, value=value)
            cell.fill = header_fill
            column_letter = cell.column_letter
            worksheet.column_dimensions[column_letter].width = col_width[col-1]
        items = [
            'Consortium data standard',
            'Funding consortium ',
            'Award number',
            'Milestone achieved',
            'Milestone completion date',
        ]
        for row, item in enumerate(items, start=2):
            worksheet.cell(row=row, column=1, value=item)
        definitions = [
            'Name of the consortium data standard under which this dataset will be processed. Examples: SPARC, HEAL',
            'SPARC, SPARC-2, VESPA, REVA, HORNET, REJOIN-HEAL, EXTERNAL',
            'Primary grant number supporting the milestone.',
            'From milestones supplied to NIH',
            'Date of milestone completion. Refer to your consortium for detailed information.'
        ]
        for row, definition in enumerate(definitions, start=2):
            worksheet.cell(row=row, column=2, value=definition)
        worksheet.cell(row=2, column=3, value='SPARC')
        worksheet.cell(row=3, column=3, value='SPARC')
        worksheet.cell(row=4, column=3,
                       value=self.__description.get_json().get('funding', ''))
        worksheet.cell(row=5, column=3, value='N/A')
        worksheet.cell(row=6, column=3, value='N/A')
        workbook.save(buffer := BytesIO())
        workbook.close()
        buffer.seek(0)
        archive.writestr(f'{filepath}', buffer.getvalue())
        buffer.close()

# ===============================================================================
